"""
App Massa - Processador de NF-e em Massa
Interface gráfica para processamento em lote de XMLs
Desenvolvido por Thucosta
"""

import os
import sys
import glob
import time
import threading
import queue
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from lxml import etree
import weasyprint
from io import BytesIO
from tkinter import ttk
import shutil
import logging
import pandas as pd
import re
import csv
from datetime import datetime
from pathlib import Path

# Configuração do CustomTkinter
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class ProcessadorMassa:
    """Classe responsável pelo processamento em massa de arquivos XML de NF-e"""
    
    def __init__(self):
        self.pasta_xmls = None
        self.pasta_saida = None
        self.template_path = None
        
        # Cache para otimização
        self.template_cache = None
        
        # Estatísticas
        self.total_arquivos = 0
        self.processados = 0
        self.sucessos = 0
        self.erros = 0
        self.inicio_processamento = None
        self.processando = False
        self.parar_solicitado = False
        
        # Dados para relatório Excel
        self.dados_relatorio = []
    
    def descobrir_xmls(self, pasta_xmls: str) -> List[str]:
        """Descobrir todos os XMLs na pasta"""
        if not os.path.exists(pasta_xmls):
            return []
        
        # Padrões de busca para XMLs de NF-e
        padroes = [
            os.path.join(pasta_xmls, "*.xml"),
            os.path.join(pasta_xmls, "**", "*.xml")
        ]
        
        xmls_encontrados = set()
        for padrao in padroes:
            xmls_encontrados.update(glob.glob(padrao, recursive=True))
        
        # Filtrar apenas arquivos válidos
        xmls_validos = [xml for xml in xmls_encontrados if os.path.isfile(xml)]
        return sorted(xmls_validos)
    
    def processar_xml_nfe(self, xml_path: str, template_content: str, output_dir: str, pdf_filename: str) -> Dict[str, Any]:
        """Processar um único XML de NF-e e gerar PDF"""
        try:
            # Ler e parsear o XML
            with open(xml_path, 'r', encoding='utf-8') as f:
                xml_content = f.read()
            
            # Parse do XML
            root = etree.fromstring(xml_content.encode('utf-8'))
            
            # Namespace da NFe
            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
            
            # Extrair dados principais da NFe
            inf_nfe = root.xpath('.//nfe:infNFe', namespaces=ns)[0]
            
            # Dados do emitente
            emit = inf_nfe.xpath('.//nfe:emit', namespaces=ns)[0]
            
            # Dados do destinatário
            dest = inf_nfe.xpath('.//nfe:dest', namespaces=ns)[0] if inf_nfe.xpath('.//nfe:dest', namespaces=ns) else None
            
            # Dados da NFe
            ide = inf_nfe.xpath('.//nfe:ide', namespaces=ns)[0]
            
            # Dados dos produtos
            produtos = inf_nfe.xpath('.//nfe:det', namespaces=ns)
            
            # Totais
            total = inf_nfe.xpath('.//nfe:total/nfe:ICMSTot', namespaces=ns)[0]
            
            # Dados de transporte
            transp = inf_nfe.xpath('.//nfe:transp', namespaces=ns)[0] if inf_nfe.xpath('.//nfe:transp', namespaces=ns) else None
            
            # Informações adicionais
            inf_adic = inf_nfe.xpath('.//nfe:infAdic', namespaces=ns)[0] if inf_nfe.xpath('.//nfe:infAdic', namespaces=ns) else None
            
            # Dados de cobrança
            cobr = inf_nfe.xpath('.//nfe:cobr', namespaces=ns)[0] if inf_nfe.xpath('.//nfe:cobr', namespaces=ns) else None
            
            # Dados para o template
            dados_nfe = {
                # Dados da NFe
                'numero': self._get_text(ide, 'nfe:nNF', ns),
                'serie': self._get_text(ide, 'nfe:serie', ns),
                'dhEmi': self._get_text(ide, 'nfe:dhEmi', ns),
                'chave': inf_nfe.get('Id', '').replace('NFe', ''),
                'natOp': self._get_text(ide, 'nfe:natOp', ns),
                
                # Emitente
                'emit_nome': self._get_text(emit, 'nfe:xNome', ns),
                'emit_cnpj': self._get_text(emit, 'nfe:CNPJ', ns),
                'emit_ie': self._get_text(emit, 'nfe:IE', ns),
                'emit_iest': self._get_text(emit, 'nfe:IEST', ns),
                'emit_endereco': self._get_endereco(emit, ns),
                
                # Destinatário
                'dest_nome': self._get_text(dest, 'nfe:xNome', ns) if dest is not None else '',
                'dest_cnpj': self._get_text(dest, 'nfe:CNPJ', ns) if dest is not None else '',
                'dest_cpf': self._get_text(dest, 'nfe:CPF', ns) if dest is not None else '',
                'dest_ie': self._get_text(dest, 'nfe:IE', ns) if dest is not None else '',
                'dest_endereco': self._get_endereco(dest, ns) if dest is not None else {},
                
                # Produtos
                'produtos': self._processar_produtos(produtos, ns),
                
                # Totais
                'vBC': self._get_text(total, 'nfe:vBC', ns),
                'vICMS': self._get_text(total, 'nfe:vICMS', ns),
                'vBCST': self._get_text(total, 'nfe:vBCST', ns),
                'vST': self._get_text(total, 'nfe:vST', ns),
                'vProd': self._get_text(total, 'nfe:vProd', ns),
                'vFrete': self._get_text(total, 'nfe:vFrete', ns),
                'vSeg': self._get_text(total, 'nfe:vSeg', ns),
                'vDesc': self._get_text(total, 'nfe:vDesc', ns),
                'vOutro': self._get_text(total, 'nfe:vOutro', ns),
                'vIPI': self._get_text(total, 'nfe:vIPI', ns),
                'vNF': self._get_text(total, 'nfe:vNF', ns),
                'vFCP': self._get_text(total, 'nfe:vFCP', ns),
                'vTotTrib': self._get_text(total, 'nfe:vTotTrib', ns),
                
                # Transporte
                'transp_dados': self._extrair_transporte(transp, ns) if transp is not None else {},
                
                # Informações adicionais
                'inf_compl': self._get_text(inf_adic, 'nfe:infCpl', ns) if inf_adic is not None else '',
                
                # Duplicatas/Fatura
                'duplicatas': self._extrair_duplicatas(cobr, ns) if cobr is not None else [],
                
                # Protocolo (se existir)
                'protocolo': self._extrair_protocolo(root, ns),
            }
            
            # Substituir variáveis no template
            html_final = self._substituir_variaveis(template_content, dados_nfe)
            
            # Gerar PDF
            pdf_path = os.path.join(output_dir, pdf_filename)
            html_doc = weasyprint.HTML(string=html_final)
            html_doc.write_pdf(pdf_path)
            
            return {
                'success': True,
                'pdf_path': pdf_path,
                'dados': dados_nfe
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _get_text(self, element, xpath: str, ns: dict) -> str:
        """Extrair texto de um elemento XML"""
        if element is None:
            return ''
        try:
            result = element.xpath(xpath, namespaces=ns)
            return result[0].text if result and result[0].text else ''
        except:
            return ''
    
    def _get_endereco(self, element, ns: dict) -> dict:
        """Extrair dados do endereço"""
        if element is None:
            return {}
        
        ender = element.xpath('.//nfe:enderEmit | .//nfe:enderDest', namespaces=ns)
        if not ender:
            return {}
        
        ender = ender[0]
        return {
            'logradouro': self._get_text(ender, 'nfe:xLgr', ns),
            'numero': self._get_text(ender, 'nfe:nro', ns),
            'complemento': self._get_text(ender, 'nfe:xCpl', ns),
            'bairro': self._get_text(ender, 'nfe:xBairro', ns),
            'cidade': self._get_text(ender, 'nfe:xMun', ns),
            'uf': self._get_text(ender, 'nfe:UF', ns),
            'cep': self._get_text(ender, 'nfe:CEP', ns),
            'fone': self._get_text(ender, 'nfe:fone', ns)
        }
    
    def _processar_produtos(self, produtos, ns: dict) -> List[Dict]:
        """Processar lista de produtos"""
        lista_produtos = []
        
        for produto in produtos:
            prod = produto.xpath('.//nfe:prod', namespaces=ns)[0]
            
            # Impostos
            imposto = produto.xpath('.//nfe:imposto', namespaces=ns)[0] if produto.xpath('.//nfe:imposto', namespaces=ns) else None
            icms_data = {}
            ipi_data = {}
            
            if imposto:
                # ICMS
                icms = imposto.xpath('.//nfe:ICMS', namespaces=ns)
                if icms:
                    icms_det = icms[0].xpath('.//*[local-name()="vBC" or local-name()="pICMS" or local-name()="vICMS"]', namespaces=ns)
                    if icms_det:
                        icms_parent = icms[0].xpath('.//*[local-name()="vBC"]', namespaces=ns)
                        if icms_parent:
                            icms_parent = icms_parent[0].getparent()
                            icms_data = {
                                'vbc': self._get_text(icms_parent, 'nfe:vBC', ns),
                                'picms': self._get_text(icms_parent, 'nfe:pICMS', ns),
                                'vicms': self._get_text(icms_parent, 'nfe:vICMS', ns)
                            }
                
                # IPI
                ipi = imposto.xpath('.//nfe:IPI', namespaces=ns)
                if ipi:
                    ipi_trib = ipi[0].xpath('.//nfe:IPITrib', namespaces=ns)
                    if ipi_trib:
                        ipi_data = {
                            'pipi': self._get_text(ipi_trib[0], 'nfe:pIPI', ns),
                            'vipi': self._get_text(ipi_trib[0], 'nfe:vIPI', ns)
                        }
            
            produto_data = {
                'codigo': self._get_text(prod, 'nfe:cProd', ns),
                'descricao': self._get_text(prod, 'nfe:xProd', ns),
                'ncm': self._get_text(prod, 'nfe:NCM', ns),
                'cfop': self._get_text(prod, 'nfe:CFOP', ns),
                'unidade': self._get_text(prod, 'nfe:uCom', ns),
                'quantidade': self._get_text(prod, 'nfe:qCom', ns),
                'valor_unitario': self._get_text(prod, 'nfe:vUnCom', ns),
                'valor_total': self._get_text(prod, 'nfe:vProd', ns),
                'icms': icms_data,
                'ipi': ipi_data
            }
            
            lista_produtos.append(produto_data)
        
        return lista_produtos
    
    def _substituir_variaveis(self, template: str, dados: Dict) -> str:
        """Substituir variáveis no template HTML"""
        html = template
        
        # Formatar datas
        dhEmi = dados.get('dhEmi', '')
        if dhEmi:
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(dhEmi.replace('T', ' ').replace('-03:00', ''))
                data_emissao = dt.strftime('%d/%m/%Y')
                hora_emissao = dt.strftime('%H:%M:%S')
            except:
                data_emissao = dhEmi[:10] if len(dhEmi) >= 10 else dhEmi
                hora_emissao = dhEmi[11:19] if len(dhEmi) >= 19 else ''
        else:
            data_emissao = ''
            hora_emissao = ''
        
        # Endereco emitente e destinatário
        emit_end = dados.get('emit_endereco', {})
        dest_end = dados.get('dest_endereco', {})
        transp_dados = dados.get('transp_dados', {})
        transporta = transp_dados.get('transporta', {})
        veiculo = transp_dados.get('veiculo', {})
        vol = transp_dados.get('vol', {})
        
        # Substituições baseadas no template real - COMPLETAS
        substituicoes = {
            # === DADOS DA EMPRESA EMITENTE ===
            '[ds_company_issuer_name]': dados.get('emit_nome', ''),
            '[ds_company_address]': f"{emit_end.get('logradouro', '')}, {emit_end.get('numero', '')}".strip(', '),
            '[ds_company_neighborhood]': emit_end.get('bairro', ''),
            '[nu_company_cep]': self._formatar_cep(emit_end.get('cep', '')),
            '[ds_company_city_name]': emit_end.get('cidade', ''),
            '[ds_company_uf]': emit_end.get('uf', ''),
            '[nl_company_phone_number]': emit_end.get('fone', ''),
            '[nl_company_cnpj_cpf]': self._formatar_cnpj_cpf(dados.get('emit_cnpj', '')),
            '[nl_company_ie]': dados.get('emit_ie', ''),
            '[nl_company_ie_st]': dados.get('emit_iest', ''),
            
            # === DADOS DA NF-E ===
            '[nl_invoice]': dados.get('numero', ''),
            '[ds_invoice_serie]': dados.get('serie', ''),
            '[ds_danfe]': dados.get('chave', ''),
            '[dt_invoice_issue]': data_emissao,
            '[dt_input_output]': data_emissao,
            '[hr_input_output]': hora_emissao,
            '[ds_code_operation_type]': '1',  # Default saída
            '[actual_page]': '1',
            '[total_pages]': '1',
            
            # === DADOS DO DESTINATÁRIO ===
            '[ds_client_receiver_name]': dados.get('dest_nome', ''),
            '[nl_client_cnpj_cpf]': self._formatar_cnpj_cpf(dados.get('dest_cnpj', '') or dados.get('dest_cpf', '')),
            '[ds_client_address]': f"{dest_end.get('logradouro', '')}, {dest_end.get('numero', '')}".strip(', '),
            '[ds_client_neighborhood]': dest_end.get('bairro', ''),
            '[nu_client_cep]': self._formatar_cep(dest_end.get('cep', '')),
            '[ds_client_city_name]': dest_end.get('cidade', ''),
            '[ds_client_uf]': dest_end.get('uf', ''),
            '[nl_client_phone_number]': dest_end.get('fone', ''),
            '[ds_client_ie]': dados.get('dest_ie', ''),
            
            # === NATUREZA DA OPERAÇÃO ===
            '[_ds_transaction_nature]': dados.get('natOp', ''),
            
            # === PROTOCOLO ===
            '[ds_protocol]': dados.get('protocolo', ''),
            '[protocol_label]': 'PROTOCOLO DE AUTORIZAÇÃO DE USO',
            
            # === CÁLCULO DO IMPOSTO - PRIMEIRA LINHA ===
            '[tot_bc_icms]': self._formatar_valor(dados.get('vBC', '')),
            '[tot_icms]': self._formatar_valor(dados.get('vICMS', '')),
            '[tot_bc_icms_st]': self._formatar_valor(dados.get('vBCST', '')),
            '[tot_icms_st]': self._formatar_valor(dados.get('vST', '')),
            '[tot_icms_fcp]': self._formatar_valor(dados.get('vFCP', '')),
            '[vl_total_prod]': self._formatar_valor(dados.get('vProd', '')),
            
            # === CÁLCULO DO IMPOSTO - SEGUNDA LINHA ===
            '[vl_shipping]': self._formatar_valor(dados.get('vFrete', '')),
            '[vl_insurance]': self._formatar_valor(dados.get('vSeg', '')),
            '[vl_discount]': self._formatar_valor(dados.get('vDesc', '')),
            '[vl_other_expense]': self._formatar_valor(dados.get('vOutro', '')),
            '[tot_total_ipi_tax]': self._formatar_valor(dados.get('vIPI', '')),
            '[vl_total]': self._formatar_valor(dados.get('vNF', '')),
            
            # === TRANSPORTADOR ===
            '[ds_transport_carrier_name]': transporta.get('nome', ''),
            '[ds_transport_code_shipping_type]': transp_dados.get('mod_frete', ''),
            '[ds_transport_rntc]': veiculo.get('rntc', ''),
            '[ds_transport_vehicle_plate]': veiculo.get('placa', ''),
            '[ds_transport_vehicle_uf]': veiculo.get('uf', ''),
            '[nl_transport_cnpj_cpf]': self._formatar_cnpj_cpf(transporta.get('cnpj', '') or transporta.get('cpf', '')),
            '[ds_transport_address]': transporta.get('endereco', ''),
            '[ds_transport_city]': transporta.get('cidade', ''),
            '[ds_transport_uf]': transporta.get('uf', ''),
            '[ds_transport_ie]': transporta.get('ie', ''),
            
            # === VOLUMES TRANSPORTADOS ===
            '[nu_transport_amount_transported_volumes]': vol.get('qvol', ''),
            '[ds_transport_type_volumes_transported]': vol.get('esp', ''),
            '[ds_transport_mark_volumes_transported]': vol.get('marca', ''),
            '[ds_transport_number_volumes_transported]': vol.get('nvol', ''),
            '[vl_transport_gross_weight]': self._formatar_valor(vol.get('peso_bruto', '')),
            '[vl_transport_net_weight]': self._formatar_valor(vol.get('peso_liquido', '')),
            
            # === INFORMAÇÕES ADICIONAIS ===
            '[ds_additional_information]': dados.get('inf_compl', ''),
            
            # === OUTROS ===
            '[barcode_image]': '',
            '{ApproximateTax}': self._formatar_valor(dados.get('vTotTrib', '')),
        }
        
        for variavel, valor in substituicoes.items():
            html = html.replace(variavel, str(valor))
        
        # === PROCESSAR DUPLICATAS ===
        duplicatas_html = ''
        duplicatas = dados.get('duplicatas', [])
        if duplicatas:
            duplicatas_html = '<table cellpadding="0" cellspacing="0" border="1" style="width: 100%;">'
            duplicatas_html += '<tr><th>Número</th><th>Vencimento</th><th>Valor</th></tr>'
            for dup in duplicatas:
                venc = self._formatar_data(dup.get('vencimento', ''))
                valor = self._formatar_valor(dup.get('valor', ''))
                duplicatas_html += f'<tr><td>{dup.get("numero", "")}</td><td>{venc}</td><td>{valor}</td></tr>'
            duplicatas_html += '</table>'
        html = html.replace('[duplicates]', duplicatas_html)
        
        # === PROCESSAR PRODUTOS ===
        produtos_html = ''
        produtos = dados.get('produtos', [])
        for produto in produtos:
            icms = produto.get('icms', {})
            ipi = produto.get('ipi', {})
            
            produtos_html += f'''
            <tr>
                <td style="text-align: center; padding: 2px;">{produto.get('codigo', '')}</td>
                <td style="padding: 2px;">{produto.get('descricao', '')}</td>
                <td style="text-align: center; padding: 2px;">{produto.get('ncm', '')}</td>
                <td style="text-align: center; padding: 2px;">{produto.get('cfop', '')}</td>
                <td style="text-align: center; padding: 2px;">{produto.get('unidade', '')}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_quantidade(produto.get('quantidade', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_valor(produto.get('valor_unitario', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_valor(produto.get('valor_total', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_valor(icms.get('vbc', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_valor(icms.get('vicms', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_valor(ipi.get('vipi', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_porcentagem(icms.get('picms', ''))}</td>
                <td style="text-align: right; padding: 2px;">{self._formatar_porcentagem(ipi.get('pipi', ''))}</td>
            </tr>
            '''
        html = html.replace('[items]', produtos_html)
        
        return html
    
    def _extrair_protocolo(self, root, ns: dict) -> str:
        """Extrair protocolo de autorização"""
        try:
            prot_nfe = root.xpath('.//nfe:protNFe', namespaces=ns)
            if prot_nfe:
                inf_prot = prot_nfe[0].xpath('.//nfe:infProt', namespaces=ns)
                if inf_prot:
                    protocolo = self._get_text(inf_prot[0], 'nfe:nProt', ns)
                    data_prot = self._get_text(inf_prot[0], 'nfe:dhRecbto', ns)
                    if protocolo and data_prot:
                        return f"{protocolo} - {data_prot[:19]}"
                    return protocolo
        except:
            pass
        return ''
    
    def _extrair_transporte(self, transp, ns: dict) -> dict:
        """Extrair dados de transporte"""
        try:
            dados_transp = {
                'mod_frete': self._get_text(transp, 'nfe:modFrete', ns),
                'transporta': {},
                'vol': {}
            }
            
            # Dados da transportadora
            transporta = transp.xpath('.//nfe:transporta', namespaces=ns)
            if transporta:
                transporta = transporta[0]
                dados_transp['transporta'] = {
                    'cnpj': self._get_text(transporta, 'nfe:CNPJ', ns),
                    'cpf': self._get_text(transporta, 'nfe:CPF', ns),
                    'nome': self._get_text(transporta, 'nfe:xNome', ns),
                    'ie': self._get_text(transporta, 'nfe:IE', ns),
                    'endereco': self._get_text(transporta, 'nfe:xEnder', ns),
                    'cidade': self._get_text(transporta, 'nfe:xMun', ns),
                    'uf': self._get_text(transporta, 'nfe:UF', ns)
                }
            
            # Dados do veículo
            veiculo = transp.xpath('.//nfe:veicTransp', namespaces=ns)
            if veiculo:
                veiculo = veiculo[0]
                dados_transp['veiculo'] = {
                    'placa': self._get_text(veiculo, 'nfe:placa', ns),
                    'uf': self._get_text(veiculo, 'nfe:UF', ns),
                    'rntc': self._get_text(veiculo, 'nfe:RNTC', ns)
                }
            
            # Volume transportado
            vol = transp.xpath('.//nfe:vol', namespaces=ns)
            if vol:
                vol = vol[0]
                dados_transp['vol'] = {
                    'qvol': self._get_text(vol, 'nfe:qVol', ns),
                    'esp': self._get_text(vol, 'nfe:esp', ns),
                    'marca': self._get_text(vol, 'nfe:marca', ns),
                    'nvol': self._get_text(vol, 'nfe:nVol', ns),
                    'peso_liquido': self._get_text(vol, 'nfe:pesoL', ns),
                    'peso_bruto': self._get_text(vol, 'nfe:pesoB', ns)
                }
            
            return dados_transp
        except:
            return {}
    
    def _extrair_duplicatas(self, cobr, ns: dict) -> list:
        """Extrair duplicatas/fatura"""
        try:
            duplicatas = []
            dups = cobr.xpath('.//nfe:dup', namespaces=ns)
            for dup in dups:
                duplicata = {
                    'numero': self._get_text(dup, 'nfe:nDup', ns),
                    'vencimento': self._get_text(dup, 'nfe:dVenc', ns),
                    'valor': self._get_text(dup, 'nfe:vDup', ns)
                }
                duplicatas.append(duplicata)
            return duplicatas
        except:
            return []
    
    def _formatar_cnpj_cpf(self, numero: str) -> str:
        """Formatar CNPJ ou CPF"""
        if not numero:
            return ''
        numero = ''.join(filter(str.isdigit, numero))
        if len(numero) == 14:  # CNPJ
            return f"{numero[:2]}.{numero[2:5]}.{numero[5:8]}/{numero[8:12]}-{numero[12:]}"
        elif len(numero) == 11:  # CPF
            return f"{numero[:3]}.{numero[3:6]}.{numero[6:9]}-{numero[9:]}"
        return numero
    
    def _formatar_cep(self, cep: str) -> str:
        """Formatar CEP"""
        if not cep:
            return ''
        cep = ''.join(filter(str.isdigit, cep))
        if len(cep) == 8:
            return f"{cep[:5]}-{cep[5:]}"
        return cep
    
    def _formatar_valor(self, valor: str) -> str:
        """Formatar valor monetário"""
        if not valor:
            return '0,00'
        try:
            num = float(valor.replace(',', '.'))
            return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return valor
    
    def _formatar_quantidade(self, qtd: str) -> str:
        """Formatar quantidade"""
        if not qtd:
            return '0,0000'
        try:
            num = float(qtd.replace(',', '.'))
            return f"{num:,.4f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return qtd
    
    def _formatar_porcentagem(self, perc: str) -> str:
        """Formatar porcentagem"""
        if not perc:
            return '0,00'
        try:
            num = float(perc.replace(',', '.'))
            return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return perc
    
    def _formatar_data(self, data: str) -> str:
        """Formatar data"""
        if not data:
            return ''
        try:
            # Se vier no formato AAAA-MM-DD
            if len(data) >= 10 and '-' in data:
                partes = data[:10].split('-')
                return f"{partes[2]}/{partes[1]}/{partes[0]}"
            return data
        except:
            return data

class NFeStudioPro(ctk.CTk):
    """NFe Studio Pro - Suite Completa para Processamento de Notas Fiscais Eletrônicas"""
    
    def __init__(self):
        super().__init__()
        
        # Configuração da janela
        self.title("🏢 NFe Studio Pro - Conversor XML→PDF e Renomeador Inteligente | Thucosta")
        self.geometry("1400x900")
        self.configure(fg_color=("#f0f0f0", "#2b2b2b"))
        
        # === VARIÁVEIS DO CONVERSOR ===
        self.pasta_xmls_var = tk.StringVar()
        self.pasta_saida_var = tk.StringVar()
        self.template_var = tk.StringVar()
        
        # Estado do processamento do conversor
        self.processando = False
        self.message_queue = queue.Queue()
        self.current_screen = "converter"
        
        # Processador do conversor
        self.processador = ProcessadorMassa()
        
        # === VARIÁVEIS DO RENOMEADOR ===
        self.dados_df = pd.DataFrame(columns=['Chave Acesso NF', 'Nome Arq. NF', 'Status'])
        self.selected_folder_rename = tk.StringVar()
        self.filtro_var_rename = tk.StringVar()
        self.status_filtro_ativo = ""
        self.filtro_var_rename.trace('w', self.aplicar_filtro_rename)
        self.status_var_rename = tk.StringVar(value="Pronto para processar arquivos")
        
        # Configurar interface
        self.setup_ui()
        
        # Iniciar verificação de mensagens do conversor apenas
        self.after(100, self.check_message_queue_converter)
    
    def create_renomeador_screen(self):
        """Criar tela do renomeador de arquivos"""
        self.renomeador_frame = ctk.CTkScrollableFrame(self.screen_container, corner_radius=0)
        
        # Header do renomeador
        self.create_renomeador_header()
        
        # Toolbar
        self.create_renomeador_toolbar()
        
        # Filter bar
        self.create_renomeador_filter_bar()
        
        # Data frame (principal)
        self.create_renomeador_data_frame()
        
        # Status bar
        self.create_renomeador_status_bar()
        
        # Setup styles para treeview
        self.setup_renomeador_styles()
        
        # Carregar dados iniciais
        self.carregar_dados_na_tree()
    
    def create_renomeador_header(self):
        """Criar cabeçalho do renomeador"""
        header_frame = ctk.CTkFrame(self.renomeador_frame, height=80, corner_radius=15)
        header_frame.pack(fill="x", pady=(0, 10))
        header_frame.pack_propagate(False)
        
        title_label = ctk.CTkLabel(
            header_frame, 
            text="🔄 Renomeador Inteligente - XML e PDF em Massa",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=20)
    
    def create_renomeador_toolbar(self):
        """Criar barra de ferramentas do renomeador"""
        self.toolbar_frame = ctk.CTkFrame(self.renomeador_frame, corner_radius=15)
        self.toolbar_frame.pack(pady=(0, 10), fill="x")
        
        # Frame principal para botões
        buttons_frame = ctk.CTkFrame(self.toolbar_frame)
        buttons_frame.pack(fill="x", pady=10)
        
        # Botões principais - organizados por função
        self.btn_pasta_rename = ctk.CTkButton(
            buttons_frame, 
            text="📁 Selecionar Pasta", 
            command=self.selecionar_pasta_rename,
            width=140, height=35
        )
        self.btn_pasta_rename.pack(side="left", padx=5)
        
        # Separador visual
        separator1 = ctk.CTkLabel(buttons_frame, text="|", text_color="gray")
        separator1.pack(side="left", padx=5)
        
        self.btn_adicionar_dados = ctk.CTkButton(
            buttons_frame, 
            text="➕ Adicionar Dados", 
            command=self.adicionar_lote_rename,
            width=140, height=35,
            fg_color="#2196F3"
        )
        self.btn_adicionar_dados.pack(side="left", padx=5)
        
        # Separador visual
        separator2 = ctk.CTkLabel(buttons_frame, text="|", text_color="gray")
        separator2.pack(side="left", padx=5)
        
        self.btn_validar = ctk.CTkButton(
            buttons_frame, 
            text="✅ Validar Tudo", 
            command=self.validar_todos_rename,
            width=140, height=35,
            fg_color="#FF9800"
        )
        self.btn_validar.pack(side="left", padx=5)
        
        # Separador visual
        separator3 = ctk.CTkLabel(buttons_frame, text="|", text_color="gray")
        separator3.pack(side="left", padx=5)
        
        self.btn_renomear = ctk.CTkButton(
            buttons_frame, 
            text="🔄 Renomear Tudo", 
            command=self.iniciar_renomeacao,
            width=140, height=35,
            fg_color="#2B8B3D",
            hover_color="#228B22"
        )
        self.btn_renomear.pack(side="left", padx=5)
        
        # Separador visual
        separator4 = ctk.CTkLabel(buttons_frame, text="|", text_color="gray")
        separator4.pack(side="left", padx=5)
        
        self.btn_limpar = ctk.CTkButton(
            buttons_frame, 
            text="🗑️ Limpar Tudo", 
            command=self.limpar_lista_rename,
            width=140, height=35,
            fg_color="#f44336"
        )
        self.btn_limpar.pack(side="left", padx=5)
        
        # Estatísticas no lado direito
        self.stats_frame = ctk.CTkFrame(buttons_frame)
        self.stats_frame.pack(side="right", padx=10)
        
        self.pasta_label = ctk.CTkLabel(
            self.stats_frame, 
            text="Nenhuma pasta selecionada",
            font=ctk.CTkFont(size=11)
        )
        self.pasta_label.pack(side="top", padx=10, pady=2)
        
        self.stats_label = ctk.CTkLabel(
            self.stats_frame, 
            text="Total: 0 | Válidos: 0 | Erros: 0",
            font=ctk.CTkFont(size=11, weight="bold")
        )
        self.stats_label.pack(side="top", padx=10, pady=2)
    
    def create_renomeador_filter_bar(self):
        """Criar barra de filtros do renomeador"""
        filter_frame = ctk.CTkFrame(self.renomeador_frame, corner_radius=15)
        filter_frame.pack(pady=(0, 10), fill="x")
        
        # Filtro por texto
        filter_label = ctk.CTkLabel(filter_frame, text="🔍 Filtrar:", font=ctk.CTkFont(size=12))
        filter_label.pack(side="left", padx=(10, 5), pady=10)
        
        self.filter_entry = ctk.CTkEntry(
            filter_frame, 
            placeholder_text="Digite para filtrar...",
            width=200,
            textvariable=self.filtro_var_rename
        )
        self.filter_entry.pack(side="left", padx=5, pady=10)
        self.filter_entry.bind('<KeyRelease>', self.aplicar_filtro_rename)
        
        # Filtros por status
        status_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        status_frame.pack(side="left", padx=20, pady=10)
        
        status_buttons = [
            ("Todos", ""),
            ("Válidos", "Válido"),
            ("Erros", "Erro"),
            ("Sucessos", "Sucesso")
        ]
        
        for text, status in status_buttons:
            btn = ctk.CTkButton(
                status_frame,
                text=text,
                command=lambda s=text: self.filtrar_por_status_rename(s),
                width=80, height=30,
                fg_color="gray70"
            )
            btn.pack(side="left", padx=2)
        
        # Ações no lado direito
        actions_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        actions_frame.pack(side="right", padx=5, pady=10)
        
        self.btn_selecionar_todos = ctk.CTkButton(
            actions_frame, text="☑️ Selecionar Todos", command=self.selecionar_todos_rename,
            width=130, height=32
        )
        self.btn_selecionar_todos.pack(side="right", padx=5)
    
    def create_renomeador_data_frame(self):
        """Criar frame principal com a tabela de dados"""
        self.data_frame = ctk.CTkFrame(self.renomeador_frame, corner_radius=15)
        self.data_frame.pack(pady=(0, 10), fill="both", expand=True)
        
        # Frame para a tabela (lado esquerdo)
        table_frame = ctk.CTkFrame(self.data_frame)
        table_frame.pack(side="left", fill="both", expand=True, padx=(10, 5), pady=10)
        
        # Configuração da Treeview
        columns = ("Chave Acesso NF", "Nome Arq. NF", "Status")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        # Configuração das colunas
        self.tree.heading("Chave Acesso NF", text="Chave Acesso NF")
        self.tree.heading("Nome Arq. NF", text="Nome Arq. NF")
        self.tree.heading("Status", text="Status")
        
        self.tree.column("Chave Acesso NF", width=300, minwidth=250)
        self.tree.column("Nome Arq. NF", width=400, minwidth=300)
        self.tree.column("Status", width=120, minwidth=100)
        
        # Scrollbars para a tabela
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack da treeview e scrollbars
        self.tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Frame para o log em tempo real (lado direito)
        log_frame = ctk.CTkFrame(self.data_frame)
        log_frame.pack(side="right", fill="both", padx=(5, 10), pady=10)
        
        # Título do log
        log_title = ctk.CTkLabel(
            log_frame, 
            text="📋 Log de Execução em Tempo Real",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        log_title.pack(pady=(10, 5))
        
        # Área de texto para o log do renomeador
        self.log_text_rename = ctk.CTkTextbox(
            log_frame, 
            width=350, 
            height=400,
            font=ctk.CTkFont(family="Consolas", size=10),
            wrap="word"
        )
        self.log_text_rename.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Botões do log
        log_buttons_frame = ctk.CTkFrame(log_frame)
        log_buttons_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.btn_limpar_log_rename = ctk.CTkButton(
            log_buttons_frame,
            text="🗑️ Limpar Log",
            command=self.limpar_log_rename,
            width=100, height=30,
            fg_color="#f44336"
        )
        self.btn_limpar_log_rename.pack(side="left", padx=5)
        
        self.btn_salvar_log_rename = ctk.CTkButton(
            log_buttons_frame,
            text="💾 Salvar Log",
            command=self.salvar_log_rename,
            width=100, height=30,
            fg_color="#2196F3"
        )
        self.btn_salvar_log_rename.pack(side="right", padx=5)
        
        # Binds
        self.tree.bind("<Double-1>", self.editar_item_rapido_rename)
        self.tree.bind("<Button-3>", self.mostrar_menu_contexto_rename)
        
        # Mensagens iniciais serão adicionadas depois que a tela estiver ativa
    
    def create_renomeador_status_bar(self):
        """Criar barra de status do renomeador"""
        status_frame = ctk.CTkFrame(self.renomeador_frame, height=60, corner_radius=15)
        status_frame.pack(side="bottom", fill="x", pady=(10, 0))
        status_frame.pack_propagate(False)
        
        # Progress bar
        self.progress_bar_rename = ctk.CTkProgressBar(status_frame, width=400)
        self.progress_bar_rename.pack(side="left", padx=10, pady=15)
        self.progress_bar_rename.set(0)
        
        # Status label
        self.status_label_rename = ctk.CTkLabel(
            status_frame, 
            textvariable=self.status_var_rename,
            font=ctk.CTkFont(size=12)
        )
        self.status_label_rename.pack(side="left", padx=10, pady=15)
        
        # Contador de itens
        self.contador_label = ctk.CTkLabel(
            status_frame, 
            text="Itens: 0 | Selecionados: 0",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.contador_label.pack(side="right", padx=10, pady=15)
    
    def center_window(self):
        """Centralizar janela na tela"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def setup_ui(self):
        """Configurar elementos da interface moderna"""
        
        # Container principal
        self.main_container = ctk.CTkFrame(self, corner_radius=0)
        self.main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Navigation Bar
        self.create_navigation()
        
        # Container para as telas
        self.screen_container = ctk.CTkFrame(self.main_container, corner_radius=0)
        self.screen_container.pack(fill="both", expand=True, pady=(10, 0))
        
        # Criar ambas as telas
        self.create_converter_screen()
        
        # Mostrar tela inicial
        self.show_screen("converter")
    
    def create_navigation(self):
        """Criar barra de navegação moderna"""
        nav_frame = ctk.CTkFrame(self.main_container, height=80, corner_radius=15)
        nav_frame.pack(fill="x", pady=(0, 10))
        nav_frame.pack_propagate(False)
        
        # Logo/Título no lado esquerdo
        logo_frame = ctk.CTkFrame(nav_frame, fg_color="transparent")
        logo_frame.pack(side="left", padx=20, pady=20)
        
        logo_label = ctk.CTkLabel(
            logo_frame,
            text="🏢 NFe Studio Pro",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        logo_label.pack()
        
        version_label = ctk.CTkLabel(
            logo_frame,
            text="v1.0 - Suite Profissional by Thucosta",
            font=ctk.CTkFont(size=10),
            text_color=("gray50", "gray50")
        )
        version_label.pack()
        
        # Botões de navegação no centro
        nav_buttons_frame = ctk.CTkFrame(nav_frame, fg_color="transparent")
        nav_buttons_frame.pack(expand=True)
        
        self.btn_converter = ctk.CTkButton(
            nav_buttons_frame,
            text="📄 Conversor XML→PDF",
            command=lambda: self.show_screen("converter"),
            width=200,
            height=40,
            corner_radius=10,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=("#2fa572", "#106A43"),
            hover_color=("#26934B", "#0D5635")
        )
        self.btn_converter.pack(side="left", padx=10)
        
        self.btn_renomeador = ctk.CTkButton(
            nav_buttons_frame,
            text="🔄 Renomeador XML/PDF",
            command=lambda: self.show_screen("renomeador"),
            width=200,
            height=40,
            corner_radius=10,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=("#3498DB", "#2980B9"),
            hover_color=("#2E86AB", "#21618C")
        )
        self.btn_renomeador.pack(side="left", padx=10)
    
    def show_screen(self, screen_name):
        """Alternar entre telas"""
        self.current_screen = screen_name
        
        # Esconder todas as telas
        if hasattr(self, 'converter_frame'):
            self.converter_frame.pack_forget()
        if hasattr(self, 'renomeador_frame'):
            self.renomeador_frame.pack_forget()
        
        # Mostrar tela selecionada
        if screen_name == "converter":
            if hasattr(self, 'converter_frame'):
                self.converter_frame.pack(fill="both", expand=True)
            self.update_nav_buttons("converter")
            # Adicionar mensagem inicial do conversor se o log estiver vazio
            if hasattr(self, 'message_text') and not self.message_text.get("1.0", "end").strip():
                self.add_message("🚀 Processador de NF-e em Massa v1.0 inicializado!")
                self.add_message("📁 Selecione a pasta com XMLs para começar o processamento")
                self.add_message("💡 Dica: Use a pasta de saída padrão ou escolha uma personalizada")
                
        elif screen_name == "renomeador":
            # Criar tela do renomeador se não existe
            if not hasattr(self, 'renomeador_frame'):
                self.create_renomeador_screen()
            self.renomeador_frame.pack(fill="both", expand=True)
            self.update_nav_buttons("renomeador")
            # Adicionar mensagem inicial do renomeador se o log estiver vazio
            if hasattr(self, 'log_text_rename') and not self.log_text_rename.get("1.0", "end").strip():
                self.adicionar_log_rename("🚀 Renomeador iniciado - Pronto para processar arquivos XML e PDF")
                self.adicionar_log_rename("💡 Selecione uma pasta e adicione dados para começar")
    
    def update_nav_buttons(self, active_screen):
        """Atualizar aparência dos botões de navegação"""
        if active_screen == "converter":
            self.btn_converter.configure(fg_color=("#2fa572", "#106A43"))
            self.btn_renomeador.configure(fg_color=("gray70", "gray30"))
        else:
            self.btn_converter.configure(fg_color=("gray70", "gray30"))
            self.btn_renomeador.configure(fg_color=("#3498DB", "#2980B9"))
    
    def create_converter_screen(self):
        """Criar tela do conversor"""
        self.converter_frame = ctk.CTkScrollableFrame(self.screen_container, corner_radius=0)
        
        # Header
        self.create_header()
        
        # Cards organizados
        self.create_config_card()
        self.create_controls_card()
        self.create_progress_card()
        self.create_log_card()
    
    def create_header(self):
        """Criar cabeçalho moderno"""
        header_frame = ctk.CTkFrame(self.converter_frame, height=120, corner_radius=15)
        header_frame.pack(fill="x", pady=(0, 20))
        header_frame.pack_propagate(False)
        
        # Título principal com ícone
        title_label = ctk.CTkLabel(
            header_frame,
            text="📄 Conversor XML→PDF Profissional",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        title_label.pack(pady=15)
        
        # Subtítulo
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Transforme XMLs NFe em DANFEs profissionais",
            font=ctk.CTkFont(size=14),
            text_color=("gray60", "gray40")
        )
        subtitle_label.pack()
        
        # Versão
        version_label = ctk.CTkLabel(
            header_frame,
            text="v1.0 - NFe Studio Pro by Thucosta",
            font=ctk.CTkFont(size=11),
            text_color=("gray50", "gray50")
        )
        version_label.pack(pady=(5, 0))
        
    def create_config_card(self):
        """Criar card de configurações moderno"""
        config_frame = ctk.CTkFrame(self.converter_frame, corner_radius=15)
        config_frame.pack(fill="x", pady=(0, 15))
        
        # Título do card
        config_title = ctk.CTkLabel(
            config_frame,
            text="📁 Configurações",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        config_title.pack(pady=(20, 15))
        
        # Container para os campos
        fields_container = ctk.CTkFrame(config_frame, fg_color="transparent")
        fields_container.pack(padx=25, pady=(0, 20), fill="x")
        
        # Campo Pasta XMLs
        self.create_input_field(
            fields_container,
            "📂 Pasta com XMLs",
            "Selecione a pasta contendo os arquivos XML das NF-e...",
            self.pasta_xmls_var,
            self.select_xmls_folder,
            "Procurar"
        )
        
        # Campo Pasta Saída  
        self.create_input_field(
            fields_container,
            "📤 Pasta de Saída",
            "Onde os PDFs serão salvos...",
            self.pasta_saida_var,
            self.select_output_folder,
            "Procurar"
        )
        
        # Campo Template
        self.create_input_field(
            fields_container,
            "🎨 Template HTML",
            "Arquivo de template para geração da DANFE...",
            self.template_var,
            self.select_template_file,
            "Procurar"
        )
    
    def create_input_field(self, parent, label, placeholder, variable, command, btn_text):
        """Criar campo de entrada moderno"""
        field_frame = ctk.CTkFrame(parent, fg_color="transparent")
        field_frame.pack(fill="x", pady=(0, 15))
        
        # Label
        label_widget = ctk.CTkLabel(
            field_frame,
            text=label,
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w"
        )
        label_widget.pack(anchor="w", pady=(0, 8))
        
        # Container do input
        input_container = ctk.CTkFrame(field_frame, fg_color="transparent")
        input_container.pack(fill="x")
        
        # Entry
        entry = ctk.CTkEntry(
            input_container,
            textvariable=variable,
            placeholder_text=placeholder,
            height=40,
            font=ctk.CTkFont(size=12),
            corner_radius=8
        )
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        # Button
        btn = ctk.CTkButton(
            input_container,
            text=btn_text,
            command=command,
            width=120,
            height=40,
            corner_radius=8,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        btn.pack(side="right")
        
        return entry
        

    def create_controls_card(self):
        """Criar card de controles moderno"""
        controls_frame = ctk.CTkFrame(self.converter_frame, corner_radius=15)
        controls_frame.pack(fill="x", pady=(0, 15))
        
        # Título do card
        controls_title = ctk.CTkLabel(
            controls_frame,
            text="⚡ Controles",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        controls_title.pack(pady=(20, 15))
        
        # Container dos botões
        buttons_container = ctk.CTkFrame(controls_frame, fg_color="transparent")
        buttons_container.pack(padx=25, pady=(0, 20), fill="x")
        
        # Botão Iniciar
        self.start_btn = ctk.CTkButton(
            buttons_container,
            text="🚀 Iniciar Processamento",
            command=self.iniciar_processamento,
            height=55,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color=("#2fa572", "#106A43"),
            hover_color=("#26934B", "#0D5635"),
            corner_radius=12
        )
        self.start_btn.pack(side="left", padx=(0, 10), expand=True, fill="x")
        
        # Botão Parar
        self.stop_btn = ctk.CTkButton(
            buttons_container,
            text="⏹️ Parar Processamento",
            command=self.parar_processamento,
            height=55,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color=("#E74C3C", "#C0392B"),
            hover_color=("#CB4335", "#A93226"),
            corner_radius=12,
            state="disabled",
            width=200
        )
        self.stop_btn.pack(side="right")
        
    def create_progress_card(self):
        """Criar card de progresso moderno"""
        progress_frame = ctk.CTkFrame(self.converter_frame, corner_radius=15)
        progress_frame.pack(fill="x", pady=(0, 15))
        
        # Título do card
        progress_title = ctk.CTkLabel(
            progress_frame,
            text="📊 Progresso do Processamento",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        progress_title.pack(pady=(20, 15))
        
        # Container do progresso
        progress_container = ctk.CTkFrame(progress_frame, fg_color="transparent")
        progress_container.pack(padx=25, pady=(0, 20), fill="x")
        
        # Barra de progresso moderna
        self.progress_bar = ctk.CTkProgressBar(
            progress_container, 
            height=25,
            corner_radius=12,
            progress_color=("#2fa572", "#106A43")
        )
        self.progress_bar.pack(fill="x", pady=(0, 15))
        self.progress_bar.set(0)
        
        # Estatísticas com design moderno
        stats_frame = ctk.CTkFrame(progress_container, corner_radius=10)
        stats_frame.pack(fill="x")
        
        self.stats_label = ctk.CTkLabel(
            stats_frame,
            text="⏳ Aguardando início do processamento...",
            font=ctk.CTkFont(size=14),
            text_color=("gray70", "gray30")
        )
        self.stats_label.pack(pady=15)
        
    def create_log_card(self):
        """Criar card de log moderno"""
        log_frame = ctk.CTkFrame(self.converter_frame, corner_radius=15)
        log_frame.pack(fill="both", expand=True)
        
        # Título do card
        log_title = ctk.CTkLabel(
            log_frame,
            text="📝 Log de Atividades",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=("#1f538d", "#14375e")
        )
        log_title.pack(pady=(20, 15))
        
        # Container do log
        log_container = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_container.pack(padx=25, pady=(0, 20), fill="both", expand=True)
        
        # Textbox do log com design moderno
        self.message_text = ctk.CTkTextbox(
            log_container,
            height=200,
            corner_radius=10,
            font=ctk.CTkFont(size=11),
            wrap="word"
        )
        self.message_text.pack(fill="both", expand=True)
        
        # Mensagens iniciais serão adicionadas quando a tela for ativada
    

    
    def select_xmls_folder(self):
        """Selecionar pasta com XMLs"""
        folder = filedialog.askdirectory(title="Selecionar pasta com XMLs de NF-e")
        if folder:
            self.pasta_xmls_var.set(folder)
            self.verificar_xmls()
    
    def select_output_folder(self):
        """Selecionar pasta de saída"""
        folder = filedialog.askdirectory(title="Selecionar pasta de saída")
        if folder:
            self.pasta_saida_var.set(folder)
    
    def select_template_file(self):
        """Selecionar arquivo de template"""
        file_path = filedialog.askopenfilename(
            title="Selecionar template HTML",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")]
        )
        if file_path:
            self.template_var.set(file_path)
    
    def verificar_xmls(self):
        """Verificar quantos XMLs existem na pasta selecionada"""
        if not self.pasta_xmls_var.get():
            return
        
        try:
            xmls = self.processador.descobrir_xmls(self.pasta_xmls_var.get())
            count = len(xmls)
            self.add_message(f"📊 Encontrados {count:,} XMLs na pasta selecionada")
            
        except Exception as e:
            self.add_message(f"❌ Erro ao verificar XMLs: {str(e)}", is_error=True)
    
    def add_message(self, message: str, is_error: bool = False):
        """Adicionar mensagem ao log do conversor (apenas se a tela conversor está ativa)"""
        # VERIFICAÇÃO RIGOROSA: Só adicionar se ESTIVER na tela do conversor E o widget do conversor existir
        if (hasattr(self, 'message_text') and 
            hasattr(self, 'current_screen') and 
            self.current_screen == "converter" and
            hasattr(self, 'converter_frame')):
            
            timestamp = datetime.now().strftime("%H:%M:%S")
            formatted_message = f"[{timestamp}] [CONVERSOR] {message}\n"
            
            self.message_text.insert("end", formatted_message)
            self.message_text.see("end")
            
            if is_error:
                # Destacar erros em vermelho (se possível)
                pass
    
    def iniciar_processamento(self):
        """Iniciar processamento em massa"""
        # Validações
        if not self.pasta_xmls_var.get():
            messagebox.showerror("Erro", "Selecione a pasta com XMLs!")
            return
        
        if not os.path.exists(self.pasta_xmls_var.get()):
            messagebox.showerror("Erro", "Pasta de XMLs não existe!")
            return
        
        if not os.path.exists(self.template_var.get()):
            messagebox.showerror("Erro", "Template HTML não encontrado!")
            return
        
        # Descobrir XMLs
        xmls = self.processador.descobrir_xmls(self.pasta_xmls_var.get())
        if not xmls:
            messagebox.showerror("Erro", "Nenhum XML encontrado na pasta!")
            return
        
        # Confirmar processamento com interface melhorada
        resposta = messagebox.askyesno(
            "🚀 Confirmar Processamento em Massa",
            f"📊 Total de XMLs encontrados: {len(xmls):,}\n"
            f"📁 Pasta de origem: {self.pasta_xmls_var.get()}\n"
            f"📤 Pasta de destino: {self.pasta_saida_var.get()}\n\n"
            f"⏱️ Tempo estimado: {len(xmls)*2:.0f}-{len(xmls)*5:.0f} segundos\n"
            f"🔧 Processamento sequencial para máxima estabilidade\n\n"
            f"Deseja iniciar o processamento?",
            icon='question'
        )
        
        if not resposta:
            return
        
        # Iniciar processamento
        self.processando = True
        self.processador.parar_solicitado = False
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        
        # Executar em thread separada
        thread = threading.Thread(target=self.executar_processamento, args=(xmls,), daemon=True)
        thread.start()
    
    def executar_processamento(self, xmls: List[str]):
        """Executar processamento em massa (roda em thread separada)"""
        try:
            # Configurar processador
            self.processador.pasta_xmls = self.pasta_xmls_var.get()
            self.processador.pasta_saida = self.pasta_saida_var.get()
            self.processador.template_path = self.template_var.get()
            self.processador.total_arquivos = len(xmls)
            self.processador.processados = 0
            self.processador.sucessos = 0
            self.processador.erros = 0
            self.processador.inicio_processamento = time.time()
            
            # Limpar dados do relatório anterior
            self.processador.dados_relatorio = []
            
            # Criar pasta de saída
            os.makedirs(self.processador.pasta_saida, exist_ok=True)
            
            # ⚡ OTIMIZAÇÃO: Carregar template uma única vez
            if self.processador.template_cache is None:
                try:
                    with open(self.processador.template_path, 'r', encoding='utf-8') as f:
                        self.processador.template_cache = f.read()
                    self.message_queue.put(("message", "⚡ Template carregado em cache"))
                except Exception as e:
                    self.message_queue.put(("message", f"❌ Erro ao carregar template: {e}"))
                    return
            
            self.message_queue.put(("message", f"🚀 Iniciando processamento de {len(xmls):,} XMLs"))
            self.message_queue.put(("message", f"📁 Pasta de saída: {self.processador.pasta_saida}"))
            self.message_queue.put(("message", "⚡ Modo otimizado: template em cache + processador reutilizado"))
            
            # Processar cada XML sequencialmente
            for i, xml_path in enumerate(xmls):
                if self.processador.parar_solicitado:
                    self.message_queue.put(("message", "⚠️ Processamento interrompido pelo usuário"))
                    break
                
                try:
                    # Nome do arquivo PDF baseado no XML
                    nome_base = Path(xml_path).stem
                    pdf_filename = f"{nome_base}.pdf"
                    pdf_path = os.path.join(self.processador.pasta_saida, pdf_filename)
                    
                    # Processar NF-e gerando apenas PDF (com cache)
                    resposta = self.processador.processar_xml_nfe(
                        xml_path=xml_path,
                        template_content=self.processador.template_cache,
                        output_dir=self.processador.pasta_saida,
                        pdf_filename=pdf_filename
                    )
                    
                    self.processador.processados += 1
                    
                    # Coletar dados para o relatório Excel
                    chave_acesso = ""
                    numero_nf = ""
                    sucesso_conversao = "Não"
                    
                    if resposta.get('success', False):
                        self.processador.sucessos += 1
                        sucesso_conversao = "Sim"
                        dados_nfe = resposta.get('dados', {})
                        chave_acesso = dados_nfe.get('chave', '')
                        numero_nf = dados_nfe.get('numero', '')
                    else:
                        self.processador.erros += 1
                        erro_msg = resposta.get('error', 'Erro desconhecido')
                        self.message_queue.put(("message", f"❌ {os.path.basename(xml_path)}: {erro_msg}"))
                        # Tentar extrair chave e número mesmo com erro
                        try:
                            with open(xml_path, 'r', encoding='utf-8') as f:
                                xml_content = f.read()
                            root = etree.fromstring(xml_content.encode('utf-8'))
                            ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                            inf_nfe = root.xpath('.//nfe:infNFe', namespaces=ns)[0]
                            chave_acesso = inf_nfe.get('Id', '').replace('NFe', '')
                            ide = inf_nfe.xpath('.//nfe:ide', namespaces=ns)[0]
                            numero_nf = ide.xpath('.//nfe:nNF', namespaces=ns)[0].text if ide.xpath('.//nfe:nNF', namespaces=ns) else ''
                        except:
                            pass
                    
                    # Adicionar dados ao relatório
                    self.processador.dados_relatorio.append({
                        'Chave de Acesso': chave_acesso,
                        'Nota Fiscal': numero_nf,
                        'Sucesso de Conversão': sucesso_conversao,
                        'Arquivo XML': os.path.basename(xml_path),
                        'Data/Hora Processamento': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                        'Pasta Origem': os.path.dirname(xml_path),
                        'Tamanho Arquivo (KB)': round(os.path.getsize(xml_path) / 1024, 2) if os.path.exists(xml_path) else 0,
                        'Erro Detalhado': resposta.get('error', '') if not resposta.get('success', False) else ''
                    })
                    
                    # Calcular progresso
                    progresso = self.processador.processados / self.processador.total_arquivos
                    tempo_decorrido = time.time() - self.processador.inicio_processamento
                    velocidade = self.processador.processados / tempo_decorrido if tempo_decorrido > 0 else 0
                    tempo_restante = (self.processador.total_arquivos - self.processador.processados) / velocidade if velocidade > 0 else 0
                    
                    # Enviar atualização de progresso
                    self.message_queue.put(("progress", {
                        'valor': progresso,
                        'processados': self.processador.processados,
                        'total': self.processador.total_arquivos,
                        'sucessos': self.processador.sucessos,
                        'erros': self.processador.erros,
                        'velocidade': velocidade,
                        'tempo_restante': tempo_restante
                    }))
                    
                    # Log otimizado: menos frequente para massa
                    if self.processador.processados % max(50, self.processador.total_arquivos // 20) == 0:
                        self.message_queue.put(("message", f"📊 Processados: {self.processador.processados}/{self.processador.total_arquivos} ({progresso*100:.1f}%)"))
                    
                except Exception as e:
                    self.processador.erros += 1
                    self.message_queue.put(("message", f"❌ Erro em {os.path.basename(xml_path)}: {str(e)}"))
                    # Adicionar ao relatório mesmo com erro crítico
                    self.processador.dados_relatorio.append({
                        'Chave de Acesso': '',
                        'Nota Fiscal': '',
                        'Sucesso de Conversão': 'Não',
                        'Arquivo XML': os.path.basename(xml_path),
                        'Data/Hora Processamento': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                        'Pasta Origem': os.path.dirname(xml_path),
                        'Tamanho Arquivo (KB)': round(os.path.getsize(xml_path) / 1024, 2) if os.path.exists(xml_path) else 0,
                        'Erro Detalhado': str(e)
                    })
            
            # Estatísticas finais
            tempo_total = time.time() - self.processador.inicio_processamento
            velocidade_media = self.processador.total_arquivos / tempo_total if tempo_total > 0 else 0
            
            # Gerar relatório Excel
            try:
                if self.processador.dados_relatorio:
                    df = pd.DataFrame(self.processador.dados_relatorio)
                    
                    # Nome do arquivo Excel com timestamp
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_filename = f"Relatorio_Conversao_NFe_{timestamp}.xlsx"
                    excel_path = os.path.join(self.processador.pasta_saida, excel_filename)
                    
                    # Salvar Excel com formatação avançada
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        # Aba principal com dados
                        df.to_excel(writer, sheet_name='Relatório Conversão', index=False)
                        
                        # Aba de estatísticas
                        sucessos = len(df[df['Sucesso de Conversão'] == 'Sim'])
                        erros = len(df[df['Sucesso de Conversão'] == 'Não'])
                        total = len(df)
                        tamanho_total = df['Tamanho Arquivo (KB)'].sum()
                        tempo_processamento = time.time() - self.processador.inicio_processamento
                        
                        stats_data = {
                            'Estatística': [
                                'Total de Arquivos',
                                'Conversões Bem-sucedidas',
                                'Conversões com Erro',
                                'Taxa de Sucesso (%)',
                                'Tamanho Total Processado (MB)',
                                'Tempo Total de Processamento (min)',
                                'Velocidade Média (arquivos/min)',
                                'Data/Hora Início',
                                'Data/Hora Fim'
                            ],
                            'Valor': [
                                total,
                                sucessos,
                                erros,
                                round((sucessos/total)*100, 2) if total > 0 else 0,
                                round(tamanho_total / 1024, 2),
                                round(tempo_processamento / 60, 2),
                                round((total / tempo_processamento) * 60, 2) if tempo_processamento > 0 else 0,
                                datetime.fromtimestamp(self.processador.inicio_processamento).strftime('%d/%m/%Y %H:%M:%S'),
                                datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                            ]
                        }
                        
                        stats_df = pd.DataFrame(stats_data)
                        stats_df.to_excel(writer, sheet_name='Estatísticas', index=False)
                        
                        # Formatação da aba principal
                        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        
                        worksheet = writer.sheets['Relatório Conversão']
                        
                        # Formatação do cabeçalho
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Formatação condicional por status
                        success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        
                        # Aplicar cores baseadas no status de conversão
                        for row in range(2, len(df) + 2):
                            status_cell = worksheet[f'C{row}']  # Coluna 'Sucesso de Conversão'
                            if status_cell.value == 'Sim':
                                for col in range(1, len(df.columns) + 1):
                                    worksheet.cell(row=row, column=col).fill = success_fill
                            elif status_cell.value == 'Não':
                                for col in range(1, len(df.columns) + 1):
                                    worksheet.cell(row=row, column=col).fill = error_fill
                        
                        # Ajustar largura das colunas
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 60)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Formatação da aba de estatísticas
                        stats_ws = writer.sheets['Estatísticas']
                        
                        # Cabeçalho das estatísticas
                        for cell in stats_ws[1]:
                            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Ajustar largura das colunas de estatísticas
                        stats_ws.column_dimensions['A'].width = 35
                        stats_ws.column_dimensions['B'].width = 25
                        
                        # Adicionar bordas
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in stats_ws.iter_rows():
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    self.message_queue.put(("message", f"📊 Relatório Excel gerado: {excel_filename}"))
                    self.message_queue.put(("excel_path", excel_path))  # Para abrir automaticamente
                    
            except Exception as e:
                self.message_queue.put(("message", f"⚠️ Erro ao gerar relatório Excel: {str(e)}"))
            
            self.message_queue.put(("message", "🎉 PROCESSAMENTO CONCLUÍDO!"))
            self.message_queue.put(("message", f"📊 Total: {self.processador.total_arquivos:,} XMLs"))
            self.message_queue.put(("message", f"✅ Sucessos: {self.processador.sucessos:,} ({self.processador.sucessos/self.processador.total_arquivos*100:.1f}%)"))
            self.message_queue.put(("message", f"❌ Erros: {self.processador.erros:,} ({self.processador.erros/self.processador.total_arquivos*100:.1f}%)"))
            self.message_queue.put(("message", f"⏱️ Tempo total: {tempo_total/60:.1f} minutos"))
            self.message_queue.put(("message", f"⚡ Velocidade média: {velocidade_media:.1f} XMLs/segundo"))
            self.message_queue.put(("message", f"📁 Arquivos salvos em: {self.processador.pasta_saida}"))
            
            self.message_queue.put(("finish", None))
            
        except Exception as e:
            self.message_queue.put(("message", f"❌ Erro geral no processamento: {str(e)}"))
            self.message_queue.put(("finish", None))
    
    def parar_processamento(self):
        """Parar processamento em massa"""
        self.processador.parar_solicitado = True
        self.processando = False
        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        
        self.add_message("⚠️ Solicitação de parada enviada...")
    
    def check_message_queue_converter(self):
        """Verificar mensagens na queue do conversor (executa na thread principal)"""
        try:
            while True:
                msg_type, msg_data = self.message_queue.get_nowait()
                
                # Só processar se estiver na tela do conversor
                if hasattr(self, 'current_screen') and self.current_screen == "converter":
                    if msg_type == "message":
                        if isinstance(msg_data, tuple):
                            message, is_error = msg_data
                            self.add_message(message, is_error)
                        else:
                            self.add_message(msg_data)
                    
                    elif msg_type == "progress":
                        # Atualizar barra de progresso se existir
                        if hasattr(self, 'progress_bar'):
                            self.progress_bar.set(msg_data['valor'])
                        
                        # Enviar estatísticas para o log do conversor (apenas ocasionalmente)
                        progresso_pct = msg_data['valor']*100
                        tempo_restante_min = msg_data['tempo_restante']/60
                        
                        # Só enviar para o log a cada 10% ou múltiplos de 50 arquivos
                        should_log = (
                            progresso_pct >= 100 or  # Sempre logar conclusão
                            progresso_pct % 10 < 0.5 or  # A cada 10%
                            msg_data['processados'] % 50 == 0  # A cada 50 arquivos
                        )
                        
                        if should_log:
                            if progresso_pct < 100:
                                # Mensagem de progresso no log
                                self.add_message(
                                    f"🔄 Progresso: {msg_data['processados']:,}/{msg_data['total']:,} "
                                    f"({progresso_pct:.1f}%) | "
                                    f"✅ Sucessos: {msg_data['sucessos']:,} | "
                                    f"❌ Erros: {msg_data['erros']:,} | "
                                    f"⚡ Velocidade: {msg_data['velocidade']:.1f} XMLs/s | "
                                    f"⏱️ Tempo restante: {tempo_restante_min:.1f} min"
                                )
                            else:
                                # Mensagem de conclusão no log
                                self.add_message(
                                    f"✅ Processamento concluído! {msg_data['processados']:,} XMLs processados | "
                                    f"Sucessos: {msg_data['sucessos']:,} | "
                                    f"Erros: {msg_data['erros']:,} | "
                                    f"⚡ Velocidade média: {msg_data['velocidade']:.1f} XMLs/s"
                                )
                        
                        # Atualizar stats_label com versão simplificada (se existir)
                        if hasattr(self, 'stats_label'):
                            stats_text_simple = (
                                f"✅ Sucessos: {msg_data['sucessos']:,} | "
                                f"❌ Erros: {msg_data['erros']:,} | "
                                f"⚡ Velocidade: {msg_data['velocidade']:.1f} XMLs/s"
                            )
                            self.stats_label.configure(text=stats_text_simple)
                    
                    elif msg_type == "excel_path":
                        # Armazenar caminho do Excel para abrir depois
                        self.excel_path_gerado = msg_data
                    
                    elif msg_type == "finish":
                        # Finalizar processamento
                        self.processando = False
                        if hasattr(self, 'start_btn'):
                            self.start_btn.configure(state="normal")
                        if hasattr(self, 'stop_btn'):
                            self.stop_btn.configure(state="disabled")
                        if hasattr(self, 'progress_bar'):
                            self.progress_bar.set(1.0)
                        
                        # Mostrar resultado final
                        excel_info = ""
                        if self.processador.dados_relatorio:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            excel_filename = f"Relatorio_Conversao_NFe_{timestamp}.xlsx"
                            excel_info = f"\n📊 Relatório Excel: {excel_filename}"
                        
                        result = messagebox.askyesnocancel(
                            "Processamento Concluído!",
                            f"Processamento finalizado com sucesso!\n\n"
                            f"✅ Sucessos: {self.processador.sucessos:,}\n"
                            f"❌ Erros: {self.processador.erros:,}\n"
                            f"📁 Arquivos salvos em:\n{self.processador.pasta_saida}"
                            f"{excel_info}\n\n"
                            f"🚀 Deseja abrir o relatório Excel agora?"
                        )
                        
                        # Se o usuário escolheu Sim (True), abrir o Excel
                        if result and hasattr(self, 'excel_path_gerado'):
                            try:
                                import subprocess
                                import platform
                                
                                if platform.system() == "Windows":
                                    os.startfile(self.excel_path_gerado)
                                elif platform.system() == "Darwin":  # macOS
                                    subprocess.call(["open", self.excel_path_gerado])
                                else:  # Linux
                                    subprocess.call(["xdg-open", self.excel_path_gerado])
                                    
                                self.add_message("📊 Relatório Excel aberto!")
                            except Exception as e:
                                self.add_message(f"⚠️ Erro ao abrir Excel: {str(e)}")
                # Se não estiver na tela do conversor, simplesmente descartar as mensagens
                        
        except queue.Empty:
            pass
        
        # Agendar próxima verificação
        self.after(100, self.check_message_queue_converter)

    # ===== MÉTODOS DO RENOMEADOR =====
    
    def setup_renomeador_styles(self):
        """Configurar estilos da Treeview"""
        style = ttk.Style()
        style.theme_use("clam")
        
        # Configurações da Treeview
        style.configure("Treeview", 
                       background="#FFFFFF",
                       foreground="#000000",
                       rowheight=25,
                       fieldbackground="#FFFFFF",
                       font=('Segoe UI', 9))
        
        style.configure("Treeview.Heading", 
                       font=('Segoe UI', 10, 'bold'),
                       background="#E1E1E1",
                       foreground="#000000")
        
        style.map("Treeview", 
                 background=[("selected", "#0078D4")],
                 foreground=[("selected", "#FFFFFF")])
        
        # Tags para status
        self.tree.tag_configure("valido", background="#D4EDDA", foreground="#155724")
        self.tree.tag_configure("erro", background="#F8D7DA", foreground="#721C24")
        self.tree.tag_configure("processando", background="#FFF3CD", foreground="#856404")
        self.tree.tag_configure("sucesso", background="#D1ECF1", foreground="#0C5460")

    def carregar_dados_na_tree(self):
        """Carregar dados do DataFrame na Treeview"""
        # Limpar tree
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Carregar dados em lotes para melhor performance
        batch_size = 100
        total_rows = len(self.dados_df)
        
        for i in range(0, total_rows, batch_size):
            batch = self.dados_df.iloc[i:i+batch_size]
            for idx, row in batch.iterrows():
                status = row['Status']
                tag = ""
                if "Válido" in status:
                    tag = "valido"
                elif "Erro" in status:
                    tag = "erro"
                elif "Processando" in status:
                    tag = "processando"
                elif "Sucesso" in status:
                    tag = "sucesso"
                    
                self.tree.insert("", "end", values=(
                    row['Chave Acesso NF'],
                    row['Nome Arq. NF'],
                    status
                ), tags=(tag,))
                
            # Atualizar interface a cada lote
            self.update_idletasks()
            
        self.atualizar_contador_rename()

    def selecionar_pasta_rename(self):
        """Selecionar pasta com os arquivos XML e PDF"""
        self.adicionar_log_rename("📁 Abrindo seletor de pasta...")
        pasta = filedialog.askdirectory(title="Selecione a pasta com os arquivos XML e PDF")
        if pasta:
            self.selected_folder_rename.set(pasta)
            pasta_nome = Path(pasta).name
            self.pasta_label.configure(text=f"📁 {pasta_nome}")
            self.status_var_rename.set(f"Pasta selecionada: {pasta_nome}")
            
            # Contar arquivos na pasta
            try:
                arquivos = os.listdir(pasta)
                xml_count = len([f for f in arquivos if f.lower().endswith('.xml')])
                pdf_count = len([f for f in arquivos if f.lower().endswith('.pdf')])
                total_arquivos = xml_count + pdf_count
                
                self.adicionar_log_rename(f"✅ Pasta selecionada: {pasta_nome}")
                self.adicionar_log_rename(f"📊 Arquivos encontrados: {total_arquivos} ({xml_count} XML, {pdf_count} PDF)")
            except Exception as e:
                self.adicionar_log_rename(f"⚠️ Erro ao contar arquivos: {str(e)}")
        else:
            self.adicionar_log_rename("❌ Seleção de pasta cancelada")

    def validar_chave_acesso_rename(self, chave: str) -> bool:
        """Validar se a chave de acesso tem formato válido"""
        if not chave:
            return False
        chave_limpa = re.sub(r'[^0-9]', '', chave)
        return len(chave_limpa) == 44

    def adicionar_lote_rename(self):
        """Adicionar múltiplas linhas de uma vez com campos separados"""
        self.adicionar_log_rename("➕ Abrindo janela de adição de dados...")
        dialog = ctk.CTkToplevel(self)
        dialog.title("Adicionar Lote de Dados")
        dialog.geometry("1100x800")
        dialog.transient(self)
        dialog.grab_set()
        
        # Centralizar
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (550)
        y = (dialog.winfo_screenheight() // 2) - (400)
        dialog.geometry(f"1100x800+{x}+{y}")
        
        # Frame principal
        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título e instruções
        title_label = ctk.CTkLabel(
            main_frame,
            text="📝 Adicionar Lote de Dados",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(pady=(0, 10))
        
        instrucoes = ctk.CTkLabel(
            main_frame,
            text="Preencha os campos abaixo. Use uma linha por registro:",
            font=ctk.CTkFont(size=12)
        )
        instrucoes.pack(pady=(0, 15))
        
        # Frame para as colunas
        columns_frame = ctk.CTkFrame(main_frame)
        columns_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        # Coluna 1 - Chave Acesso NF
        col1_frame = ctk.CTkFrame(columns_frame)
        col1_frame.pack(side="left", fill="both", expand=True, padx=(10, 5), pady=10)
        
        chave_label = ctk.CTkLabel(
            col1_frame,
            text="🔑 Chave Acesso NF",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        chave_label.pack(pady=(10, 5))
        
        chave_info = ctk.CTkLabel(
            col1_frame,
            text="(44 dígitos numéricos)",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        chave_info.pack(pady=(0, 10))
        
        chave_text = ctk.CTkTextbox(
            col1_frame, 
            height=450,
            font=ctk.CTkFont(family="Consolas", size=10),
            wrap="none"
        )
        chave_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Placeholder para chaves
        chave_text.insert("1.0", "35200500004286000183550000010171616466719332\n35200482901640001025500100053056711680170900\n35200482901640001025500100052530810541324200\n")
        
        # Coluna 2 - Nome Arq. NF
        col2_frame = ctk.CTkFrame(columns_frame)
        col2_frame.pack(side="left", fill="both", expand=True, padx=(5, 10), pady=10)
        
        nome_label = ctk.CTkLabel(
            col2_frame,
            text="📄 Nome Arq. NF",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        nome_label.pack(pady=(10, 5))
        
        nome_info = ctk.CTkLabel(
            col2_frame,
            text="(Nome desejado para o arquivo)",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        nome_info.pack(pady=(0, 10))
        
        nome_text = ctk.CTkTextbox(
            col2_frame, 
            height=450,
            font=ctk.CTkFont(family="Consolas", size=10),
            wrap="word"
        )
        nome_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Placeholder para nomes
        nome_text.insert("1.0", "NF000101716 HELCA I E E COMERCIO DE MATERIAL CIRURGICO LTDA 4500925168 01.09.2025\nNF000530567 ESTABLISHMENT LABS BRASIL PRODUTOS SAUDE LTDA 4500322253 23.07.2025\nNF000525308 ESTABLISHMENT LABS BRASIL PRODUTOS SAUDE LTDA 4500915256 17.06.2025\n")
        
        def processar_lote():
            chaves_texto = chave_text.get("1.0", "end-1c").strip()
            nomes_texto = nome_text.get("1.0", "end-1c").strip()
            
            if not chaves_texto or not nomes_texto:
                messagebox.showwarning("Aviso", "Preencha ambas as colunas")
                return
            
            self.adicionar_log_rename("🔄 Processando lote de dados...")
            
            chaves_linhas = [linha.strip() for linha in chaves_texto.splitlines() if linha.strip()]
            nomes_linhas = [linha.strip() for linha in nomes_texto.splitlines() if linha.strip()]
            
            self.adicionar_log_rename(f"📝 Chaves inseridas: {len(chaves_linhas)}")
            self.adicionar_log_rename(f"📝 Nomes inseridos: {len(nomes_linhas)}")
            
            if len(chaves_linhas) != len(nomes_linhas):
                error_msg = f"Número de linhas diferente! Chaves: {len(chaves_linhas)}, Nomes: {len(nomes_linhas)}"
                self.adicionar_log_rename(f"❌ {error_msg}")
                messagebox.showwarning("Aviso", f"Número de linhas diferente!\n\nChaves: {len(chaves_linhas)} linhas\nNomes: {len(nomes_linhas)} linhas\n\nCertifique-se de que cada chave tenha um nome correspondente.")
                return
                
            novos_dados = []
            erros = []
            
            for i, (chave, nome) in enumerate(zip(chaves_linhas, nomes_linhas), 1):
                if not chave or not nome:
                    erros.append(f"Linha {i}: Chave ou nome vazio")
                    continue
                    
                # Validar chave
                if not self.validar_chave_acesso_rename(chave):
                    status = f"Erro - Linha {i}: Chave inválida"
                    erros.append(f"Linha {i}: Chave inválida ({chave[:20]}...)")
                    self.adicionar_log_rename(f"⚠️ Linha {i}: Chave inválida - {chave[:20]}...")
                else:
                    status = "Válido"
                    self.adicionar_log_rename(f"✅ Linha {i}: Chave válida - {nome}")
                    
                novos_dados.append({
                    'Chave Acesso NF': chave,
                    'Nome Arq. NF': nome,
                    'Status': status
                })
                
            if novos_dados:
                df_novo = pd.DataFrame(novos_dados)
                self.dados_df = pd.concat([self.dados_df, df_novo], ignore_index=True)
                self.carregar_dados_na_tree()
                
                # Mostrar resultado
                resultado_msg = f"✅ Adicionados {len(novos_dados)} registros em lote"
                self.adicionar_log_rename(f"✅ Lote processado: {len(novos_dados)} registros adicionados")
                
                if erros:
                    self.adicionar_log_rename(f"⚠️ {len(erros)} erros encontrados no lote")
                    resultado_msg += f"\n\n⚠️ {len(erros)} erros encontrados:\n" + "\n".join(erros[:5])
                    if len(erros) > 5:
                        resultado_msg += f"\n... e mais {len(erros) - 5} erros"
                
                dialog.destroy()
                messagebox.showinfo("Lote Processado", resultado_msg)
                self.status_var_rename.set(f"✅ Lote adicionado: {len(novos_dados)} itens")
            else:
                self.adicionar_log_rename("❌ Nenhum dado válido encontrado no lote")
                messagebox.showwarning("Aviso", "Nenhum dado válido encontrado")
                
        def limpar_campos():
            chave_text.delete("1.0", "end")
            nome_text.delete("1.0", "end")
            self.adicionar_log_rename("🧹 Campos limpos")
        
        # Botões
        buttons_frame = ctk.CTkFrame(main_frame)
        buttons_frame.pack(fill="x", pady=(0, 10))
        
        limpar_btn = ctk.CTkButton(
            buttons_frame,
            text="🧹 Limpar",
            command=limpar_campos,
            width=120,
            height=35
        )
        limpar_btn.pack(side="left", padx=(10, 5))
        
        cancelar_btn = ctk.CTkButton(
            buttons_frame,
            text="❌ Cancelar",
            command=dialog.destroy,
            width=120,
            height=35
        )
        cancelar_btn.pack(side="right", padx=(5, 10))
        
        processar_btn = ctk.CTkButton(
            buttons_frame,
            text="✅ Processar Lote",
            command=processar_lote,
            width=150,
            height=35
        )
        processar_btn.pack(side="right", padx=(5, 5))

    def validar_todos_rename(self):
        """Validar todas as chaves de acesso"""
        if self.dados_df.empty:
            messagebox.showwarning("Aviso", "Nenhum dado para validar")
            return
            
        self.adicionar_log_rename("🔄 Iniciando validação em massa...")
        
        # Criar barra de progresso
        progress_dialog = ctk.CTkToplevel(self)
        progress_dialog.title("Validando Dados...")
        progress_dialog.geometry("400x150")
        progress_dialog.transient(self)
        progress_dialog.grab_set()
        
        # Centralizar
        progress_dialog.update_idletasks()
        x = (progress_dialog.winfo_screenwidth() // 2) - 200
        y = (progress_dialog.winfo_screenheight() // 2) - 75
        progress_dialog.geometry(f"400x150+{x}+{y}")
        
        # Labels e progress bar
        status_label = ctk.CTkLabel(progress_dialog, text="Validando chaves de acesso...")
        status_label.pack(pady=20)
        
        progress_bar = ctk.CTkProgressBar(progress_dialog, width=350)
        progress_bar.pack(pady=10)
        progress_bar.set(0)
        
        progress_text = ctk.CTkLabel(progress_dialog, text="0%")
        progress_text.pack(pady=5)
        
        total_rows = len(self.dados_df)
        validados = 0
        erros = 0
        
        # Processar em lotes
        batch_size = 50
        for i in range(0, total_rows, batch_size):
            batch_end = min(i + batch_size, total_rows)
            
            for idx in range(i, batch_end):
                chave = str(self.dados_df.iloc[idx]['Chave Acesso NF'])
                
                if self.validar_chave_acesso_rename(chave):
                    self.dados_df.loc[idx, 'Status'] = "Válido"
                    validados += 1
                else:
                    self.dados_df.loc[idx, 'Status'] = "Erro - Chave inválida"
                    erros += 1
                    
                # Atualizar progress
                progress = (idx + 1) / total_rows
                progress_bar.set(progress)
                progress_text.configure(text=f"{int(progress * 100)}%")
                status_label.configure(text=f"Validando... {idx + 1}/{total_rows}")
                
            # Atualizar interface
            progress_dialog.update()
            self.update_idletasks()
        
        # Recarregar dados na tree
        self.carregar_dados_na_tree()
        
        # Fechar dialog
        progress_dialog.destroy()
        
        # Mostrar resultado
        resultado_msg = f"✅ Validação concluída!\n\n"
        resultado_msg += f"📊 Total processado: {total_rows}\n"
        resultado_msg += f"✅ Válidos: {validados}\n"
        resultado_msg += f"❌ Erros: {erros}"
        
        self.adicionar_log_rename(f"✅ Validação concluída: {validados} válidos, {erros} erros")
        self.status_var_rename.set(f"✅ Validação: {validados} válidos, {erros} erros")
        
        messagebox.showinfo("Validação Concluída", resultado_msg)

    def aplicar_filtro_rename(self, *args):
        """Aplicar filtro de texto em tempo real"""
        if not hasattr(self, 'tree') or not hasattr(self, 'filtro_var_rename'):
            return
            
        filtro_texto = self.filtro_var_rename.get().lower()
        
        # Limpar tree
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Filtrar dados
        dados_filtrados = self.dados_df.copy()
        
        if filtro_texto:
            # Filtrar por chave ou nome
            mascara = (
                dados_filtrados['Chave Acesso NF'].astype(str).str.lower().str.contains(filtro_texto, na=False) |
                dados_filtrados['Nome Arq. NF'].astype(str).str.lower().str.contains(filtro_texto, na=False)
            )
            dados_filtrados = dados_filtrados[mascara]
        
        # Aplicar filtro de status se existir
        if hasattr(self, 'status_filtro_ativo') and self.status_filtro_ativo != "Todos":
            if self.status_filtro_ativo == "Válidos":
                dados_filtrados = dados_filtrados[dados_filtrados['Status'] == "Válido"]
            elif self.status_filtro_ativo == "Erros":
                dados_filtrados = dados_filtrados[dados_filtrados['Status'].str.contains("Erro", na=False)]
        
        # Carregar dados filtrados
        for idx, row in dados_filtrados.iterrows():
            status = row['Status']
            tag = ""
            if "Válido" in status:
                tag = "valido"
            elif "Erro" in status:
                tag = "erro"
            elif "Processando" in status:
                tag = "processando"
            elif "Sucesso" in status:
                tag = "sucesso"
                
            self.tree.insert("", "end", values=(
                row['Chave Acesso NF'],
                row['Nome Arq. NF'],
                status
            ), tags=(tag,))
        
        # Atualizar contador
        total_filtrado = len(dados_filtrados)
        total_geral = len(self.dados_df)
        self.status_var_rename.set(f"Mostrando {total_filtrado} de {total_geral} itens")

    def filtrar_por_status_rename(self, status_filtro: str):
        """Filtrar por status específico"""
        self.status_filtro_ativo = status_filtro
        self.adicionar_log_rename(f"🔍 Filtro aplicado: {status_filtro}")
        self.aplicar_filtro_rename()

    def iniciar_renomeacao(self):
        """Iniciar processo de renomeação"""
        if self.dados_df.empty:
            messagebox.showwarning("Aviso", "Nenhum dado para processar")
            return
            
        if not self.selected_folder_rename.get():
            messagebox.showwarning("Aviso", "Selecione uma pasta primeiro")
            return
        
        # Verificar se há dados válidos
        dados_validos = self.dados_df[self.dados_df['Status'] == "Válido"]
        if dados_validos.empty:
            messagebox.showwarning("Aviso", "Nenhum dado válido para renomear")
            return
        
        # Confirmação
        total_validos = len(dados_validos)
        confirmacao = messagebox.askyesno(
            "Confirmar Renomeação",
            f"Deseja renomear {total_validos} arquivos?\n\n"
            f"📁 Pasta: {Path(self.selected_folder_rename.get()).name}\n"
            f"📄 Arquivos válidos: {total_validos}\n\n"
            "⚠️ Esta operação não pode ser desfeita!"
        )
        
        if not confirmacao:
            self.adicionar_log_rename("❌ Renomeação cancelada pelo usuário")
            return
        
        self.adicionar_log_rename("🚀 Iniciando processo de renomeação...")
        
        # Desabilitar botões durante processamento
        if hasattr(self, 'renomear_btn'):
            self.renomear_btn.configure(state="disabled", text="🔄 Renomeando...")
        
        # Iniciar thread de renomeação
        import threading
        self.renomeacao_thread = threading.Thread(target=self.renomear_arquivos_thread, daemon=True)
        self.renomeacao_thread.start()

    def limpar_lista_rename(self):
        """Limpar lista de dados"""
        self.dados_df = pd.DataFrame(columns=['Chave Acesso NF', 'Nome Arq. NF', 'Status'])
        if hasattr(self, 'tree'):
            self.carregar_dados_na_tree()

    def atualizar_contador_rename(self):
        """Atualizar contador de itens"""
        if hasattr(self, 'status_var_rename'):
            total = len(self.dados_df)
            self.status_var_rename.set(f"Total: {total} itens")

    def adicionar_log_rename(self, mensagem: str):
        """Adicionar mensagem ao log do renomeador (apenas se a tela renomeador está ativa)"""
        # VERIFICAÇÃO RIGOROSA: Só adicionar se ESTIVER na tela do renomeador E o widget do renomeador existir
        if (hasattr(self, 'log_text_rename') and 
            hasattr(self, 'current_screen') and 
            self.current_screen == "renomeador" and
            hasattr(self, 'renomeador_frame')):
            
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_text_rename.insert("end", f"[{timestamp}] [RENOMEADOR] {mensagem}\n")
            self.log_text_rename.see("end")

    def limpar_log_rename(self):
        """Limpar log"""
        if hasattr(self, 'log_text_rename'):
            self.log_text_rename.delete("1.0", "end")

    def salvar_log_rename(self):
        """Salvar log em arquivo"""
        if not hasattr(self, 'log_text_rename'):
            messagebox.showwarning("Aviso", "Nenhum log para salvar")
            return
            
        log_content = self.log_text_rename.get("1.0", "end-1c")
        if not log_content.strip():
            messagebox.showwarning("Aviso", "Log vazio")
            return
        
        # Selecionar arquivo para salvar
        arquivo = filedialog.asksaveasfilename(
            title="Salvar Log",
            defaultextension=".txt",
            filetypes=[
                ("Arquivos de texto", "*.txt"),
                ("Todos os arquivos", "*.*")
            ],
            initialname=f"log_renomeador_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if arquivo:
            try:
                with open(arquivo, 'w', encoding='utf-8') as f:
                    f.write(f"=== LOG DO RENOMEADOR NFe ===\n")
                    f.write(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                    f.write(f"Total de itens: {len(self.dados_df)}\n")
                    f.write("=" * 50 + "\n\n")
                    f.write(log_content)
                    
                self.adicionar_log_rename(f"💾 Log salvo: {os.path.basename(arquivo)}")
                messagebox.showinfo("Sucesso", f"Log salvo em:\n{arquivo}")
                
            except Exception as e:
                self.adicionar_log_rename(f"❌ Erro ao salvar log: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao salvar log:\n{str(e)}")
        else:
            self.adicionar_log_rename("❌ Salvamento de log cancelado")

    def selecionar_todos_rename(self):
        """Selecionar todos os itens na TreeView"""
        if hasattr(self, 'tree'):
            for item in self.tree.get_children():
                self.tree.selection_add(item)
            self.adicionar_log_rename(f"✅ Selecionados todos os {len(self.tree.get_children())} itens")

    def editar_item_rapido_rename(self, event):
        """Editar item rapidamente com duplo clique"""
        if not hasattr(self, 'tree'):
            return
            
        item = self.tree.selection()[0] if self.tree.selection() else None
        if not item:
            return
            
        values = self.tree.item(item, 'values')
        if not values:
            return
            
        chave_acesso = values[0]
        nome_atual = values[1]
        
        # Dialog de edição
        dialog = ctk.CTkToplevel(self)
        dialog.title("Editar Item")
        dialog.geometry("500x300")
        dialog.transient(self)
        dialog.grab_set()
        
        # Centralizar
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 250
        y = (dialog.winfo_screenheight() // 2) - 150
        dialog.geometry(f"500x300+{x}+{y}")
        
        # Campos
        ctk.CTkLabel(dialog, text="🔑 Chave Acesso NF:", font=ctk.CTkFont(weight="bold")).pack(pady=(20, 5))
        chave_entry = ctk.CTkEntry(dialog, width=400, height=35)
        chave_entry.pack(pady=5, padx=20)
        chave_entry.insert(0, chave_acesso)
        
        ctk.CTkLabel(dialog, text="📄 Nome Arquivo NF:", font=ctk.CTkFont(weight="bold")).pack(pady=(15, 5))
        nome_entry = ctk.CTkTextbox(dialog, width=400, height=100)
        nome_entry.pack(pady=5, padx=20)
        nome_entry.insert("1.0", nome_atual)
        
        def salvar():
            nova_chave = chave_entry.get().strip()
            novo_nome = nome_entry.get("1.0", "end-1c").strip()
            
            if not nova_chave or not novo_nome:
                messagebox.showwarning("Aviso", "Preencha todos os campos")
                return
                
            # Atualizar DataFrame
            mask = self.dados_df['Chave Acesso NF'] == chave_acesso
            self.dados_df.loc[mask, 'Chave Acesso NF'] = nova_chave
            self.dados_df.loc[mask, 'Nome Arq. NF'] = novo_nome
            
            # Revalidar se a chave mudou
            if nova_chave != chave_acesso:
                if self.validar_chave_acesso_rename(nova_chave):
                    self.dados_df.loc[mask, 'Status'] = "Válido"
                else:
                    self.dados_df.loc[mask, 'Status'] = "Erro - Chave inválida"
            
            self.carregar_dados_na_tree()
            self.adicionar_log_rename(f"✏️ Item editado: {nova_chave[:20]}...")
            dialog.destroy()
        
        # Botões
        buttons_frame = ctk.CTkFrame(dialog)
        buttons_frame.pack(fill="x", pady=20, padx=20)
        
        ctk.CTkButton(buttons_frame, text="❌ Cancelar", command=dialog.destroy).pack(side="right", padx=(5, 0))
        ctk.CTkButton(buttons_frame, text="✅ Salvar", command=salvar).pack(side="right", padx=(5, 5))

    def mostrar_menu_contexto_rename(self, event):
        """Mostrar menu de contexto com clique direito"""
        if not hasattr(self, 'tree'):
            return
            
        item = self.tree.identify_row(event.y)
        if not item:
            return
            
        self.tree.selection_set(item)
        
        # Criar menu de contexto
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="✏️ Editar", command=lambda: self.editar_item_rapido_rename(None))
        menu.add_command(label="📋 Duplicar", command=lambda: self.duplicar_item_rename(item))
        menu.add_separator()
        menu.add_command(label="✅ Marcar como Válido", command=lambda: self.marcar_status_rename(item, "Válido"))
        menu.add_command(label="❌ Marcar como Erro", command=lambda: self.marcar_status_rename(item, "Erro - Manual"))
        menu.add_separator()
        menu.add_command(label="🗑️ Remover", command=lambda: self.remover_item_rename(item))
        
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def remover_item_rename(self, item):
        """Remover item da lista"""
        if not hasattr(self, 'tree') or not item:
            return
            
        values = self.tree.item(item, 'values')
        if not values:
            return
            
        chave_acesso = values[0]
        
        confirmacao = messagebox.askyesno("Confirmar Remoção", f"Remover item?\n\nChave: {chave_acesso}")
        if confirmacao:
            # Remover do DataFrame
            self.dados_df = self.dados_df[self.dados_df['Chave Acesso NF'] != chave_acesso]
            self.carregar_dados_na_tree()
            self.adicionar_log_rename(f"🗑️ Item removido: {chave_acesso[:20]}...")

    def duplicar_item_rename(self, item):
        """Duplicar item selecionado"""
        if not hasattr(self, 'tree') or not item:
            return
            
        values = self.tree.item(item, 'values')
        if not values:
            return
            
        chave_acesso = values[0]
        nome_arquivo = values[1]
        
        # Criar novo item duplicado
        novo_item = {
            'Chave Acesso NF': chave_acesso + "_COPIA",
            'Nome Arq. NF': nome_arquivo + " (Cópia)",
            'Status': "Erro - Chave inválida"
        }
        
        df_novo = pd.DataFrame([novo_item])
        self.dados_df = pd.concat([self.dados_df, df_novo], ignore_index=True)
        self.carregar_dados_na_tree()
        self.adicionar_log_rename(f"📋 Item duplicado: {chave_acesso[:20]}...")

    def marcar_status_rename(self, item, novo_status: str):
        """Marcar status específico do item"""
        if not hasattr(self, 'tree') or not item:
            return
            
        values = self.tree.item(item, 'values')
        if not values:
            return
            
        chave_acesso = values[0]
        
        # Atualizar status
        self.dados_df.loc[self.dados_df['Chave Acesso NF'] == chave_acesso, 'Status'] = novo_status
        self.carregar_dados_na_tree()
        self.adicionar_log_rename(f"🏷️ Status alterado para '{novo_status}': {chave_acesso[:20]}...")

    def renomear_arquivos_thread(self):
        """Thread para renomeação de arquivos"""
        pasta = self.selected_folder_rename.get()
        dados_validos = self.dados_df[self.dados_df['Status'] == "Válido"].copy()
        
        total_arquivos = len(dados_validos)
        arquivos_renomeados = 0
        arquivos_nao_encontrados = 0
        erros = 0
        
        try:
            # Listar todos os arquivos da pasta
            arquivos_pasta = os.listdir(pasta)
            
            for idx, (_, row) in enumerate(dados_validos.iterrows()):
                chave_acesso = str(row['Chave Acesso NF'])
                novo_nome = str(row['Nome Arq. NF'])
                
                # Atualizar status para processando
                self.dados_df.loc[self.dados_df['Chave Acesso NF'] == chave_acesso, 'Status'] = "Processando..."
                
                # Atualizar interface na thread principal
                self.after(0, self.carregar_dados_na_tree)
                self.after(0, lambda i=idx+1, t=total_arquivos: self.adicionar_log_rename(f"🔄 Processando {i}/{t}: {chave_acesso[:20]}..."))
                
                # Procurar arquivo com essa chave
                arquivo_encontrado = None
                for arquivo in arquivos_pasta:
                    if chave_acesso in arquivo:
                        arquivo_encontrado = arquivo
                        break
                
                if not arquivo_encontrado:
                    # Arquivo não encontrado
                    self.dados_df.loc[self.dados_df['Chave Acesso NF'] == chave_acesso, 'Status'] = "Erro - Arquivo não encontrado"
                    arquivos_nao_encontrados += 1
                    self.after(0, lambda: self.adicionar_log_rename(f"❌ Arquivo não encontrado para chave: {chave_acesso[:20]}..."))
                    continue
                
                try:
                    # Construir caminhos
                    arquivo_original = os.path.join(pasta, arquivo_encontrado)
                    extensao = os.path.splitext(arquivo_encontrado)[1]
                    
                    # Limpar nome do arquivo (remover caracteres inválidos)
                    nome_limpo = re.sub(r'[<>:"/\\|?*]', '', novo_nome)
                    nome_limpo = nome_limpo.strip()
                    
                    if not nome_limpo:
                        nome_limpo = f"NF_{chave_acesso}"
                    
                    arquivo_novo = os.path.join(pasta, f"{nome_limpo}{extensao}")
                    
                    # Verificar se arquivo de destino já existe
                    contador = 1
                    arquivo_final = arquivo_novo
                    while os.path.exists(arquivo_final):
                        nome_com_contador = f"{nome_limpo} ({contador})"
                        arquivo_final = os.path.join(pasta, f"{nome_com_contador}{extensao}")
                        contador += 1
                    
                    # Renomear arquivo
                    shutil.move(arquivo_original, arquivo_final)
                    
                    # Atualizar status
                    nome_final = os.path.basename(arquivo_final)
                    self.dados_df.loc[self.dados_df['Chave Acesso NF'] == chave_acesso, 'Status'] = f"Sucesso - {nome_final}"
                    arquivos_renomeados += 1
                    
                    self.after(0, lambda nome=nome_final: self.adicionar_log_rename(f"✅ Renomeado: {nome}"))
                    
                except Exception as e:
                    # Erro durante renomeação
                    self.dados_df.loc[self.dados_df['Chave Acesso NF'] == chave_acesso, 'Status'] = f"Erro - {str(e)}"
                    erros += 1
                    self.after(0, lambda erro=str(e): self.adicionar_log_rename(f"❌ Erro: {erro}"))
                
                # Pequena pausa para não sobrecarregar
                import time
                time.sleep(0.1)
            
            # Finalizar processo
            def finalizar_renomeacao():
                # Reabilitar botões
                if hasattr(self, 'renomear_btn'):
                    self.renomear_btn.configure(state="normal", text="🔄 Renomear Tudo")
                
                # Recarregar dados
                self.carregar_dados_na_tree()
                
                # Log final
                self.adicionar_log_rename("🎉 Processo de renomeação concluído!")
                self.adicionar_log_rename(f"📊 Resumo: {arquivos_renomeados} renomeados, {arquivos_nao_encontrados} não encontrados, {erros} erros")
                
                # Mostrar resultado
                resultado_msg = f"🎉 Renomeação concluída!\n\n"
                resultado_msg += f"✅ Arquivos renomeados: {arquivos_renomeados}\n"
                resultado_msg += f"❓ Não encontrados: {arquivos_nao_encontrados}\n"
                resultado_msg += f"❌ Erros: {erros}\n"
                resultado_msg += f"📊 Total processado: {total_arquivos}"
                
                messagebox.showinfo("Renomeação Concluída", resultado_msg)
                self.status_var_rename.set(f"✅ Concluído: {arquivos_renomeados} renomeados")
            
            self.after(0, finalizar_renomeacao)
            
        except Exception as e:
            def mostrar_erro():
                if hasattr(self, 'renomear_btn'):
                    self.renomear_btn.configure(state="normal", text="🔄 Renomear Tudo")
                self.adicionar_log_rename(f"💥 Erro crítico: {str(e)}")
                messagebox.showerror("Erro", f"Erro durante renomeação:\n{str(e)}")
            
            self.after(0, mostrar_erro)


if __name__ == "__main__":
    app = NFeStudioPro()
    app.mainloop()
